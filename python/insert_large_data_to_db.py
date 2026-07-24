import mysql.connector
from openpyxl import load_workbook

# 1. Database Configuration
db_config = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "ecommerce",
}

# 2. File and Table Configuration
EXCEL_FILE = "orders.xlsx"
SHEET_NAME = "Sheet1"
TARGET_TABLE = "orders"

# 3. SQL Insert Statement (Update column names and place holders %s)
# Example matches a table with 3 columns: col1, col2, col3
INSERT_QUERY = f"INSERT INTO {TARGET_TABLE} (order_id, customer_id, product_id,order_date,quantity,payment_mode) VALUES (%s, %s, %s,%s, %s, %s)"
BATCH_SIZE = 5000  # Inserts data in blocks of 5000 to maximize speed

try:
    # Connect to MySQL
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    # 1. TEMPORARILY DISABLE FOREIGN KEY CHECKS
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
    print("Foreign key constraints temporarily disabled.")

    # Open Excel file in read-only mode to save memory
    wb = load_workbook(filename=EXCEL_FILE, read_only=True, data_only=True)
    sheet = wb[SHEET_NAME]

    batch = []
    total_rows = 0

    # Iterate through rows natively
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # Skip header row

        # Clean row data (Handling empty lines)
        if any(item is not None for item in row):
            batch.append(row)

        # Batch insert when limit is reached
        if len(batch) >= BATCH_SIZE:
            cursor.executemany(INSERT_QUERY, batch)
            conn.commit()
            total_rows += len(batch)
            print(f"Uploaded {total_rows} rows...")
            batch = []

    # Insert remaining rows
    if batch:
        cursor.executemany(INSERT_QUERY, batch)
        conn.commit()
        total_rows += len(batch)

    print(f"Success! Finished importing a total of {total_rows} rows.")

except mysql.connector.Error as err:
    print(f"Database Error: {err}")
except Exception as e:
    print(f"Error: {e}")
finally:
    if "cursor" in locals():
        cursor.close()
    if "conn" in locals():
        conn.close()
    if "wb" in locals():
        wb.close()